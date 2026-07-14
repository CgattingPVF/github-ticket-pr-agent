from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


_REPOSITORY_PATTERN = re.compile(r'^[A-Za-z0-9_.-]+/[A-Za-z0-9_.-]+$')


def _text(value: object) -> str:
    return '' if value is None else str(value).strip()


def _priority(value: object) -> str:
    value = _text(value).upper()
    return value if value in {'P0', 'P1', 'P2'} else ''


def _gh_candidates() -> list[Path]:
    candidates: list[Path] = []
    discovered = shutil.which('gh')
    if discovered:
        candidates.append(Path(discovered))

    program_files = os.getenv('ProgramFiles')
    if program_files:
        candidates.append(Path(program_files) / 'GitHub CLI' / 'gh.exe')

    local_app_data = os.getenv('LOCALAPPDATA')
    if local_app_data:
        candidates.append(Path(local_app_data) / 'Programs' / 'GitHub CLI' / 'gh.exe')

    user_profile = os.getenv('USERPROFILE')
    if user_profile:
        candidates.append(Path(user_profile) / 'AppData' / 'Local' / 'Programs' / 'GitHub CLI' / 'gh.exe')

    return candidates


def find_gh_executable() -> str:
    for candidate in _gh_candidates():
        if candidate.is_file():
            return str(candidate)
    raise RuntimeError(
        'GitHub CLI (gh) was not found. Install GitHub CLI, then run '
        '`gh auth login --hostname github.com` before reopening Ticket PR Agent.'
    )


def _run_gh(arguments: list[str], timeout: int = 60, token: str | None = None) -> subprocess.CompletedProcess[str]:
    kwargs: dict[str, object] = {
        'capture_output': True,
        'text': True,
        'timeout': timeout,
    }
    if os.name == 'nt':
        kwargs['creationflags'] = getattr(subprocess, 'CREATE_NO_WINDOW', 0)

    env = os.environ.copy()
    if token:
        env['GH_TOKEN'] = token

    try:
        return subprocess.run([find_gh_executable(), *arguments], **kwargs, env=env)
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError('GitHub CLI request timed out. Check your connection and try again.') from exc
    except OSError as exc:
        raise RuntimeError(f'GitHub CLI could not be started: {exc}') from exc


def _github_api_error(repository: str, result: subprocess.CompletedProcess[str]) -> RuntimeError:
    details = (result.stderr or result.stdout or 'GitHub API request failed').strip()
    lowered = details.lower()

    if 'not logged into any github hosts' in lowered or 'authentication' in lowered or 'bad credentials' in lowered:
        return RuntimeError(
            'GitHub CLI is not authenticated. Open PowerShell and run '
            '`gh auth login --hostname github.com`, then restart the app. '
            f'GitHub said: {details}'
        )

    if 'http 404' in lowered or 'not found' in lowered:
        target = repository or 'the requested repositories'
        return RuntimeError(
            f'GitHub could not access `{target}`. Check the owner/repository name and make sure the '
            'account shown by `gh auth status` has access to that repository. Private organisation '
            f'repositories may also require SSO authorisation. GitHub said: {details}'
        )

    if 'http 403' in lowered or 'resource not accessible' in lowered or 'forbidden' in lowered:
        return RuntimeError(
            'The authenticated GitHub account does not have permission to read these issues. '
            'Run `gh auth status`, refresh the token scopes with `gh auth refresh -h github.com -s repo`, '
            f'and authorise organisation SSO if required. GitHub said: {details}'
        )

    return RuntimeError(f'GitHub issue sync failed: {details}')


def _project_metadata(repository: str, issue_numbers: list[int], token: str | None) -> dict[int, dict[str, str]]:
    """Read GitHub Projects v2 priority and status fields in small GraphQL batches."""
    owner, name = repository.split('/', 1)
    metadata: dict[int, dict[str, str]] = {}
    for start in range(0, len(issue_numbers), 20):
        numbers = issue_numbers[start:start + 20]
        fields = ' '.join(
            f'i{number}: issue(number: {number}) {{ projectItems(first: 10) {{ nodes {{ fieldValues(first: 30) {{ nodes {{ ... on ProjectV2ItemFieldSingleSelectValue {{ name field {{ ... on ProjectV2SingleSelectField {{ name }} }} }} }} }} }} }} }}'
            for number in numbers
        )
        result = _run_gh(['api', 'graphql', '-f', f'query=query {{ repository(owner: "{owner}", name: "{name}") {{ {fields} }} }}'], token=token)
        if result.returncode:
            continue
        try:
            data = json.loads(result.stdout)['data']['repository']
            for number in numbers:
                for item in (data.get(f'i{number}') or {}).get('projectItems', {}).get('nodes', []):
                    for value in item.get('fieldValues', {}).get('nodes', []):
                        field_name = value.get('field', {}).get('name')
                        if field_name in {'Project Priority', 'Priority'}:
                            priority = _priority(value.get('name'))
                            if priority:
                                metadata.setdefault(number, {})['priority'] = priority
                        elif field_name in {'Project Status', 'Status'} and value.get('name'):
                            metadata.setdefault(number, {})['project_status'] = _text(value.get('name'))
        except (KeyError, TypeError, json.JSONDecodeError):
            continue
    return metadata


def import_workbook(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_text(value) for value in rows[0]]
    index = {name: i for i, name in enumerate(headers) if name}
    required = {'Issue_URL', 'Issue_Number', 'Issue_Title'}
    missing = required - index.keys()
    if missing:
        raise ValueError(f'Workbook is missing columns: {", ".join(sorted(missing))}')
    now = datetime.now(timezone.utc).isoformat()
    tickets = []
    for row in rows[1:]:
        get = lambda name: _text(row[index[name]]) if index[name] < len(row) else ''
        url = get('Issue_URL')
        if not url or '/issues/' not in url:
            continue
        repository = get('PF__Repository')
        if not repository:
            parts = url.split('/')
            repository = '/'.join(parts[3:5]) if len(parts) > 4 else ''
        number = int(float(get('Issue_Number'))) if get('Issue_Number') else 0
        tickets.append({'key': f'{repository}#{number}', 'repository': repository, 'number': number,
                        'url': url, 'title': get('Issue_Title'), 'state': get('Issue_State').upper() or 'OPEN',
                        'labels': get('Issue_Labels'), 'assignees': get('Issue_Assignees'),
                        'priority': _priority(get('PF__Project Priority')), 'project_status': get('PF__Status'),
                        'issue_type': get('Issue_Type'), 'created_at': get('Created_At'),
                        'updated_at': get('Updated_At'), 'synced_at': now, 'source': 'xlsx'})
    return tickets


def sync_github(repository: str = '', state: str = 'open', limit: int = 100, token: str | None = None) -> list[dict]:
    repository = repository.strip()
    if repository and not _REPOSITORY_PATTERN.fullmatch(repository):
        raise ValueError('Repository must use the `owner/repository` format.')

    endpoint = (
        f'repos/{repository}/issues?state={state}&per_page={min(limit, 100)}'
        if repository
        else f'search/issues?q=state:{state}+is:issue&per_page={min(limit, 100)}'
    )
    result = _run_gh(['api', endpoint], token=token)
    if result.returncode:
        raise _github_api_error(repository, result)

    try:
        payload = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f'GitHub returned invalid JSON: {exc}') from exc

    items = payload.get('items', []) if isinstance(payload, dict) else payload
    project_metadata = _project_metadata(repository, [issue['number'] for issue in items], token) if repository else {}
    now = datetime.now(timezone.utc).isoformat()
    tickets = []
    for issue in items:
        repo = issue.get('repository', {}).get('full_name', repository) if isinstance(issue, dict) else repository
        labels = ', '.join(label.get('name', '') for label in issue.get('labels', []))
        label_names = {label.get('name', '').upper() for label in issue.get('labels', [])}
        metadata = project_metadata.get(issue['number'], {})
        priority = metadata.get('priority') or next((value for value in ('P0', 'P1', 'P2') if value in label_names), '')
        assignees = ', '.join(person.get('login', '') for person in issue.get('assignees', []))
        tickets.append({'key': f'{repo}#{issue["number"]}', 'repository': repo, 'number': issue['number'],
                        'url': issue.get('html_url', ''), 'title': issue.get('title', ''),
                        'state': issue.get('state', 'open').upper(), 'labels': labels, 'assignees': assignees,
                        'priority': priority, 'project_status': metadata.get('project_status', ''), 'issue_type': 'Issue',
                        'created_at': issue.get('created_at', ''), 'updated_at': issue.get('updated_at', ''),
                        'synced_at': now, 'source': 'github'})
    return tickets
